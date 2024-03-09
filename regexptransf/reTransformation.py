import regex as re
import openpyxl

def build_metrics_dict(workbook):
    sheet_metrics = workbook["Metrics"]
    metrics_dict = {}
    
    for row in sheet_metrics.iter_rows():
        if row[0].value is None or row[4].value is None:
            continue  # continue to the next row instead of breaking
        key = (row[0].value or '', row[2].value or '', row[3].value or '')  # convert None to empty strings
        metrics_dict[key] = row[4].value
    return metrics_dict

def get_metric_from_memory(metrics_dict, t_value, r_value, c_value):
    if r_value == ['']:
        metric = metrics_dict.get((t_value, '', c_value))
    elif c_value == ['']:
        metric = metrics_dict.get((t_value, r_value, ''))
    else:
        metric = metrics_dict.get((t_value, r_value, c_value))
    return metric

def create_z_1_6_dict(sheet):
    z_1_6_data = {}
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=9):  # Assuming data starts at row 2
        t_value = row[0].value
        # Replace None with '' in the list comprehension
        values = [cell.value if cell.value is not None else '' for cell in row[2:8]]  # D to I
        z_1_6_data[t_value] = values
    return z_1_6_data

def create_KeyText_dict(sheet):
    KeyText_data = {}
    for row in sheet.iter_rows(min_row=2):  # Assuming data starts at row 2
        key_a = row[0].value  # Value from Column A
        key_c = row[2].value  # Value from Column C
        value_b = row[1].value  # Value from Column B

        # Replace None with an empty string
        key_a = key_a if key_a is not None else ''
        key_c = key_c if key_c is not None else ''

        # Combine A and C as the key
        key = f"{key_a}_{key_c}"

        KeyText_data[key] = value_b

    return KeyText_data

def get_c_values_from_structure_c(structure_c):
    if not structure_c or not isinstance(structure_c, str):
        return ['']
    match = re.search(r'scope\({t: .*?, c:(.*?), f: .*?, fv: .*?}\)', structure_c)
    return match.group(1).strip().split(";") if match else ['']
    
def get_r_values_from_structure_c(structure_c):
    if not structure_c or not isinstance(structure_c, str):
        return ['']
    match = re.search(r'scope\({t: .*?, r:(.*?), f: .*?, fv: .*?}\)', structure_c)
    return match.group(1).strip().split(";") if match else ['']

def compute_ps(t_value, r_value, c_value, input_str, metrics_dict):
    

    last_placeholder = ""
    metric_type = get_metric_from_memory(metrics_dict, t_value, r_value.strip(), c_value)
    suffix = ""  # default suffix

    if metric_type in ["Metric: Monetary", "Metric: Pure", "Metric: Decimal", "Metric: Integer"]:
        prefix = "VALUE_SPECIFIC"  # default prefix for these metric types
        if "not(isnull(" in input_str.lower():
            prefix = "( EXISTS" if input_str.startswith("(") else "EXISTS"
            suffix = ") " if input_str.endswith(")))") else ""
        elif "isnull(" in input_str.lower():
            prefix = "( NOT ( EXISTS" if input_str.startswith("(") else "NOT ( EXISTS"
            suffix = ") )" if input_str.endswith("))") else ") "
    else:
        prefix = "TEXT"  # default prefix for other metric types
        if "not(isnull(" in input_str.lower():
            prefix = "( EXISTS_TEXT" if input_str.startswith("(") else "EXISTS_TEXT"
            suffix = ") " if input_str.endswith(")))") else ""
        elif "isnull(" in input_str.lower():
            prefix = "( NOT ( EXISTS_TEXT" if input_str.startswith("(") else "NOT ( EXISTS_TEXT"
            suffix = ") )" if input_str.endswith("))") else ") "
        elif "matches" in input_str:
            prefix = "LIKE( TEXT"
   
    match = re.search(r'\[s2c_([A-Za-z0-9]+:[A-Za-z0-9]+)\]', input_str)
    if match:
        last_placeholder = ", 'TXTMD'"
    elif not match and "TEXT" in prefix and ((t_value[:10] == 'S.08.01.01' or t_value[:10] == 'S.08.01.04') and c_value == 'C0360'):
        last_placeholder = ", 'TXTLG'"
    elif not match and "TEXT" in prefix:
        last_placeholder = ", 'ELEM'"
    else: ""

    

    return {
        'prefix': prefix,
        'suffix': suffix,
        'last_placeholder': last_placeholder
    }


def convert_structure(input_str, c_values_from_structure_c, r_values_from_structure_c, metrics_dict):
    
    result = input_str
    results = []
    filter_pattern = r'filter: \[s2c_dim:[A-Z]{2}\] = \[s2c_([A-Z]{2}):x(\d+)\](?: and \[s2c_dim:[A-Z]{2}\] = \[s2c_([A-Z]{2}):x(\d+)\])?}'
    modified_pattern = r',\s' + filter_pattern
    filter_match = re.search(filter_pattern, input_str)
    filter_content = ''
    filter_content2 = ''

    
    if filter_match:
        filter_content = f"{filter_match.group(1)}:x{filter_match.group(2)}"
        filter_content2 = f"{filter_match.group(3)}:x{filter_match.group(4)}" if filter_match.group(3) else ''
        input_str = re.sub(modified_pattern, '}', input_str, count=1)

    # Check for filter and handle accordingly
    

    starts_with_sum = input_str.lower().startswith('sum')
    match = re.search(r'({t: (.*?)(, r: (.*?))?(, c: ((?:[^,]+?)(?=(?:, |}))))?(, z: (.*?))?(, .*?)?})' + 
                 r'(, "\^\.\.((?:\(\w+\.\)\|?)+)\$"\)\))?', input_str)
    if not match:
        print('does not match')
        return input_str

    r_values = match.group(4).split(';') if match.group(4) else r_values_from_structure_c
    c_values = match.group(6).split(';') if match.group(6) else c_values_from_structure_c

    t_value = match.group(2)
    if t_value.startswith("SE"):
        segments = t_value.split(".")
        segments[0] = "S"
        if len(segments[2]) == 1:
            segments[2] = "0" + segments[2]
        if len(segments) > 3:
            if segments[3] == "16":
                segments[3] = "01"
            elif segments[3] == "17":
                segments[3] = "02"
        t_value = ".".join(segments)[:10]
    
    for r_value in r_values:
     if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
        r_value = r_value.replace("ER", "R", 1)
        
    
    t_value = t_value[:10]
    

    
    
    
    if "matches" in input_str:
        r_value = r_values[0] if match.group(4) else r_values_from_structure_c
        c_value = c_values[0] if match.group(6) else c_values_from_structure_c

        process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
        
        
        segments = re.split('\||}|\\(\\(|\)\|', input_str)
        pattern = r'\^(.*?)\(\('
        m1 = re.search(pattern, input_str)
        m2 = m1.group(1) if m1 else ''
        
        pattern1 = r'\)\)(.*?)\$'
        m3 = re.search(pattern1, input_str)
        m5 = m3.group(1) if m3 else ''

        

        matchres = []
        
        for s in segments[2:] if len(segments) > 2 else segments[1:]:
          
          s = s.replace(", \"^"," \"^").replace("((","").replace("(",'').replace(")",'').replace("^","").replace("$","").replace("\"","").replace(" ","").replace("\\","")
          
          matchre = f"{process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']} , '{m2}{s}{m5}')"
          matchres.append(matchre)
        
        if input_str.startswith("not"):
           return f"NOT ( {' AND '.join(matchres)} )"
        else:
           return f" {' OR '.join(matchres)} "
    
    if not starts_with_sum:
       
        for i in range(len(r_values)):
            if t_value[:10] == "S.01.01.01" and r_values[i].startswith("ER"):
                r_values[i] = r_values[i].replace("ER", "R", 1)
        
        r_value = r_values[0] if match.group(4) else r_values_from_structure_c     
        c_value = c_values[0] if match.group(6) else c_values_from_structure_c
        
        process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
        result = f"{process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']}"
        
    
    if starts_with_sum and match.group(4):
      
        if len(r_values) > 1: 
            
            
            c_value = c_values[0] if match.group(6) else c_values_from_structure_c
            results = []
            for r_value in r_values:
                if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
                  r_value = r_value.replace("ER", "R", 1)
                process_structure = compute_ps(t_value, r_value.strip(), c_value.strip(), input_str, metrics_dict)
                structure_B = f"{process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']}"
                results.append(structure_B)
                
            result =  ' + '.join(results)

        elif len(r_values) == 1:

            r_value = r_values[0]
            if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
                r_value = r_value.replace("ER", "R", 1)
            c_value = c_values[0] if match.group(6) else c_values_from_structure_c
            t_value = t_value[:10]
            process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
            condition = "1 == 1"
            if t_value.startswith('S.06.02'):
                condition = f"TEXT( #REPPER, '#COMPANY', '{t_value}', 'SCRC0150', '', '#FISCVARNT', '', '', '', '', '', '', 'ELEM' ) <> 'AGG_ONLY'"

            
            result =  f"SUM( '{t_value}', {condition}, {process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']} )"
        
        
        elif len(r_values) < 1:  # This handles the case when there's no r_value
            r_value = ''  # Or any default value
        
            c_value = c_values[0] if match.group(6) else c_values_from_structure_c
            t_value = t_value[:10]
            process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
            condition = "1 == 1"
            if t_value.startswith('S.06.02'):
                condition = f"TEXT( #REPPER, '#COMPANY', '{t_value}', 'SCRC0150', '', '#FISCVARNT', '', '', '', '', '', '', 'ELEM' ) <> 'AGG_ONLY'"

            
            result =  f"SUM( '{t_value}', {condition}, {process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']} )"
        

    if starts_with_sum and match.group(6):
        
        if len(c_values) > 1: 
            r_value = r_values[0] if match.group(4) else r_values_from_structure_c
            if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
                r_value = r_value.replace("ER", "R", 1)
            results = []
            for c_value in c_values:
                process_structure = compute_ps(t_value, r_value.strip(), c_value.strip(), input_str, metrics_dict)
                structure_B = f"{process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']}"
                results.append(structure_B)
                
            result =  ' + '.join(results)

        elif len(c_values) == 1:
            c_value = c_values[0]
            r_value = r_values[0] if match.group(4) else r_values_from_structure_c
            if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
                r_value = r_value.replace("ER", "R", 1)
            t_value = t_value[:10]
            process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
            condition = "1 == 1"
            if t_value.startswith('S.06.02'):
                condition = f"TEXT( #REPPER, '#COMPANY', '{t_value}', 'SCRC0150', '', '#FISCVARNT', '', '', '', '', '', '', 'ELEM' ) <> 'AGG_ONLY'"

            
            result =  f"SUM( '{t_value}', {condition}, {process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']} )"
        
        
        elif len(c_values) < 1:  # This handles the case when there's no r_value
            c_value = ''  # Or any default value
        
            r_value = r_values[0] if match.group(4) else r_values_from_structure_c
            if t_value[:10] == "S.01.01.01" and r_value and r_value.startswith("ER"):
                r_value = r_value.replace("ER", "R", 1)
            t_value = t_value[:10]
            process_structure = compute_ps(t_value, r_value, c_value, input_str, metrics_dict)
            condition = "1 == 1"
            if t_value.startswith('S.06.02'):
                condition = f"TEXT( #REPPER, '#COMPANY', '{t_value}', 'SCRC0150', '', '#FISCVARNT', '', '', '', '', '', '', 'ELEM' ) <> 'AGG_ONLY'"
            
            result =  f"SUM( '{t_value}', {condition}, {process_structure['prefix']}( #REPPER, '#COMPANY', '{t_value}', '{c_value}', '{r_value}', '#FISCVARNT', '', '', '', '', '', ''{process_structure['last_placeholder']} ) {process_structure['suffix']} )"
        
    z_values = []
    if filter_match:
        # Retrieve z_values, replacing None with empty string ''
        z_values = [f"{v}" if v is not None else "" for v in z_1_6_data.get(t_value[:10], ['', '', '', '', '', ''])]
        keytext = ''
        matched_index = -1 
        matched_index2 = -1
        # New code to check for combinations
        found_combination = False
        found_combination2 = False

        for i, z_value in enumerate(z_values):
            key = f"{filter_content}_{z_value}"
            keytext = keytext_data.get(key)

            if keytext is not None:  # Check if keytext is not None
                z_values[i] = keytext
                found_combination = True
                matched_index = i
                break

        if found_combination:
            
            # New loop to find second combination if first is found
            for i, z_value in enumerate(z_values):
                key2 = f"{filter_content2}_{z_value}"
                keytext2 = keytext_data.get(key2)

                if keytext2 is not None and i != matched_index:  # Check for second keytext
                    z_values[i] = keytext2
                    found_combination2 = True
                    matched_index2 = i
                    break

        # If both combinations are found, set all others to empty except the matched ones
        if found_combination and found_combination2 and keytext is not None and keytext2 is not None:
            z_values = ['' if i != matched_index and i != matched_index2 else z_values[i] for i in range(len(z_values))]
        elif found_combination and filter_content2 == '':
            # If only the first combination is found
            z_values = ['' if i != matched_index else z_values[i] for i in range(len(z_values))]
        else:
            print('No match found')

        # Split the result string to isolate parts around the last 6 '' placeholders
        result = re.sub(r"('')(, ''){5}", 
                        f"'{z_values[0]}', '{z_values[1]}', '{z_values[2]}', '{z_values[3]}', '{z_values[4]}', '{z_values[5]}'", 
                        result)


    return result

     
def add_brackets(expression):
    # List of operators to match
    operators = ['AND', 'OR', '<>', '<=', '>=', '=', '<', '>', '+', '-', '*', '/']

    
    # If the expression contains "EXISTS", return it as is
    if ("EXISTS" in expression and len(re.findall(r'\b(AND|OR)\b', expression)) > 1) or "LIKE" in expression:
        
        return expression
    
    # For each operator
    for operator in operators:
        if operator in expression:
            # Split the expression at the operator
            parts = expression.split(operator, 1)  # Split at the first occurrence of operator

            # Check if the part ends with '1' before and starts with '1' after '==' and handle it specially
            if (parts[0].strip().endswith('1') or parts[1].strip().startswith('1')) and operator == '=':
                return expression  # if it's '1 == 1', return it as is

            # Check if the part is a number and wrap it with brackets
            left = ' ( ' + parts[0].strip() + ' ) ' if not parts[0].strip().replace('.', '', 1).isdigit() else ' ( ' + parts[0].strip() + ' ) '
            right = ' ( ' + parts[1].strip() + ' ) ' if not parts[1].strip().replace('.', '', 1).isdigit() else ' ( ' + parts[1].strip() + ' ) '
            
            # Return the bracketed parts joined by the operator
            return left + operator + right

    # If no operator matched, return the original expression
    return expression

def convert_if(equation, c_values_from_structure_c, r_values_from_structure_c, workbook, metrics_dict):

    
    if "if" in equation:
        r1 = []

        if 's2c' in equation and 'if (' in equation:
            parts1 = re.split(r"if | then false\(\) else true\(\)| then true\(\) else false\(\)| then |else true\(\)|else false\(\)", equation)
        elif 'if (' in equation and ('if false() then true()' in equation or 'if true() then false()' in equation):
            parts1 = re.split(r"if \(|\) then false\(\) else true\(\)|\) then true\(\) else false\(\)", equation)
        elif 'if (' in equation:
           parts1 = re.split(r"if \(|\) then false\(\) else true\(\)|\) then true\(\) else false\(\)|\) then |else true\(\)|else false\(\)", equation) 
        elif  'if (' not in equation and ('if false() then true()' in equation or 'if true() then false()' in equation):
            parts1 = re.split(r"if | then false\(\) else true\(\)| then true\(\) else false\(\)", equation)
        else:
           parts1 = re.split(r"if | then false\(\) else true\(\)| then true\(\) else false\(\)| then |else true\(\)|else false\(\)", equation)
        
        for part1 in parts1:
          r = convert_equation(part1, c_values_from_structure_c, r_values_from_structure_c, workbook, metrics_dict)
          if any(char.isalpha() for char in r):
                r1.append(r)
        if equation.strip().endswith("then false() else true()") and "then false() else true()" in equation:
               return 'IF(' + ', '.join(r1) + ', 1 @<@ 1, 1 == 1 )' 
        if equation.strip().endswith("then true() else false()"):
               return 'IF(' + ', '.join(r1) + ', 1 == 1, 1 @<@ 1 )'   
        if equation.strip().endswith("else true()") and "then false() else true()" not in equation:
            return 'IF(' + ', '.join(r1) + ', 1 == 1 )'
        elif equation.strip().endswith("else false()"):
            return 'IF(' + ', '.join(r1) + ', 1 @<@ 1 )'
        else:
            return 'IF( ' + ', '.join(r1) + ' )'
    else:
       return convert_equation(equation, c_values_from_structure_c, r_values_from_structure_c, workbook, metrics_dict)

def find_function_expressions(expression, function_name):
    matches = []
    stack = []
    pattern = rf'\b{function_name}\s*\('

    for match in re.finditer(pattern, expression, re.IGNORECASE):
        index = match.start()
        stack.append(index)
        count = 1
        for i in range(index + len(match.group()), len(expression)):
            if expression[i] == '(':
                count += 1
            elif expression[i] == ')':
                count -= 1
            if count == 0:
                start = stack.pop()
                matches.append(expression[start:i + 1])
                break
    return matches

def reformat_expression(expression, function_name):
    matches = find_function_expressions(expression, function_name)

    if not matches:
        return expression  # No 'min' or 'max' found; return the expression unchanged.

    for match in matches:
        inner_content = match[match.index('(') + 1:-1]
        elements = [elem.strip() for elem in re.split(r',\s*(?![^\(\)]*\))', inner_content)]

        if len(elements) > 2:
            nested_expression = elements.pop()
            while elements:
                nested_expression = f'{function_name}({elements.pop()}, {nested_expression})'
            expression = expression.replace(match, nested_expression)

    return expression

def reformat_min_max_expression(expression):
    expression = reformat_expression(expression, 'min')
    expression = reformat_expression(expression, 'max')
    return expression

def convert_equation(equation, c_values_from_structure_c, r_values_from_structure_c, workbook, metrics_dict):
    
    

    if "matches" in equation:
      
      parts = re.split(r' *(and|or|(?<!matches)\((?=matches)|(?<!matches)(?<![iI][sS][nN][uU][lL][lL])\((?=\{t:)|(?<=\])\)|(?<=\]\))\)|(?<=\]\]\))\)) *', equation)
      new_parts = []
      for part in parts:
            if part is not None and 'matches' in part and 'not(matches' not in part:
                  split_parts = re.split(r'(?<=\$"\))', part)
                  new_parts.extend(split_parts)
            else:
                  new_parts.append(part)
            parts = new_parts
    elif 'not(isnull' not in equation.lower() and 'isnull' not in equation.lower() and 'sum' not in equation:
        
        if  '-1 * imax(imin' in equation:
            equation = equation.replace('-1', '(0.00 - 1.00)')
        parts = re.split(r' *(<=|>=|==|!=(?! \[s2c)|\+|-|=(?! \[s2c)|, (?! ?([a-zA-Z]))|and|or|iabs\(|abs\(|i?max\([^)]*\)\)|i?min\([^)]*\)\)|i?max\(|i?min\(|i?max[^)]+\)|i?min[^)]+\)|(?<![a-zA-Z])\(|>|<|\)\)|\) *|\*) *', equation)
        new_parts = []
        for part in parts:
                if part is not None and (('imax' in part or 'max' in part) or ('imin' in part or 'min' in part)):
                    split_parts = re.split(r' *(, (?! ?([a-zA-Z]))|\-|\(|\)) *', part)
                    new_parts.extend(split_parts)
                else:
                    new_parts.append(part)
                parts = new_parts
    elif 'sum' in equation:
        
        parts = re.split(r' *(<=|>=|==|!=(?! \[s2c)|\+|-|=(?! \[s2c)|, (?! ?([a-zA-Z]|{))|and(?!.*\[s2c_dim:)|or|iabs\(|abs\(|i?max\([^)]*\)\)|i?min\([^)]*\)\)|i?max\(|i?min\(|i?max[^)]+\)|i?min[^)]+\)|(?<![a-zA-Z])\(|>|< *|\*) *', equation)
        new_parts = []
        for part in parts:
                if part is not None and all(sub not in part.lower() for sub in ['not(isnull)', 'sum', 'matches', 'isnull']):
                    split_parts = re.split(r'(?<=\})', part)
                    new_parts.extend(split_parts)
                elif part is not None and any(sub not in part.lower() for sub in [ 'sum', 'matches', 'isnull']):
                    split_parts = re.split(r'(?<=\))', part)
                    new_parts.extend(split_parts)
                else:
                    new_parts.append(part)
                parts = new_parts

    else:
       
        parts = re.split(r' *(<=|>=|==|!=(?! \[s2c)|\+|-|=(?! \[s2c)|, (?! ?([a-zA-Z]))|and|or|iabs\(|abs\(|i?max\([^)]*\)\)|i?min\([^)]*\)\)|i?max\(|i?min\(|i?max[^)]+\)|i?min[^)]+\)|(?<![a-zA-Z])\(|>|<|\*) *', equation)

        new_parts = []
        for part in parts:
                if part is not None and 'isnull' in part.lower() and 'not(isnull' not in part.lower() and part.count('))') > 0:
                    
                    split_parts = re.split(r'(?<=\))', part)
                    new_parts.extend(split_parts)

                    additional_parts = []
                    for split_part in new_parts:
                        if split_part is not None and split_part.startswith('not(isNull') and not split_part.endswith('))'):
                            # Further split at 'not('
                            further_split_parts = re.split(r'(?<=not\()', split_part)
                            additional_parts.extend(further_split_parts)
                        else:
                            additional_parts.append(split_part)
                    new_parts = additional_parts

                elif part is not None and 'not(isnull' in part.lower() :
                   split_parts = re.split(r'(?<=\)\))', part)
                   split_parts = re.split(r'(?<=not\()(?=not)', part)
                   new_parts.extend(split_parts)
                else:
                    new_parts.append(part)
                parts = new_parts
        
    
    for i in range(len(parts)):
        if parts[i] is None:
           parts[i] = ''  # convert None to empty string
        

        pattern = r'(\s*(=|!=)\s*)\[s2c_([A-Za-z0-9]+:[A-Za-z0-9]+)\](\))?'
        match = re.search(pattern, parts[i])
        matches1 = re.findall(pattern, parts[i])
        
        if match and 'dim' not in parts[i]:
            operator = "=" if match.group(2) == "=" else "<>"
            closing_bracket = match.group(4) or ""
            parts[i] = convert_structure(parts[i], c_values_from_structure_c, r_values_from_structure_c, metrics_dict) + f" {operator} '"  + match.group(3) + "'" + closing_bracket

        elif match and 'dim' in parts[i] and len(matches1) > 1 and 'and [s2c_dim' not in parts[i]:
            # Process the second match
            second_match = re.search(pattern, parts[i], pos=match.end())
            if second_match:
                operator = "=" if second_match.group(2) == "=" else "<>"
                closing_bracket = second_match.group(4) or ""
                parts[i] = convert_structure(parts[i], c_values_from_structure_c, r_values_from_structure_c, metrics_dict) + f" {operator} '"  + second_match.group(3) + "'" + closing_bracket
        elif 'sum' in parts[i] and ', {' in parts[i]:
            subparts = re.split(r"sum\(|, (?! ?([a-zA-Z]))", parts[i])
            subparts = [subpart for subpart in subparts if subpart and subpart.strip()]  # Filter out subparts that are just whitespace
            parts[i] = '+ '.join([convert_structure(subpart, c_values_from_structure_c, r_values_from_structure_c, metrics_dict) for subpart in subparts])

        elif '{t:' in parts[i]:
            parts[i] = convert_structure(parts[i], c_values_from_structure_c, r_values_from_structure_c, metrics_dict)
        elif parts[i].lower() in ['!=', '+', '-', '=', '*', 'and', 'or', '>', '<', '<=', '>=', 'imin', 'min','imax', 'max', '\(', '\)', ' , ','iabs(', 'abs(', 'not(']:
            parts[i] = ' ' + parts[i].upper() + ' '
        elif parts[i].replace('.', '', 1).isdigit() or parts[i].replace(' ', '',).isdigit() or (
                parts[i].startswith('-') and parts[i][1:].replace('.', '', 1).isdigit()):
             if float(parts[i]) == 0:
                parts[i] = "0.00"
              # Otherwise, leave parts[i] unchanged.
    
    converted_equation = ''.join([str(x) if x is not None else '' for x in parts]).strip().replace('IMIN', 'MIN').replace('IMAX', 'MAX').replace("IABS", "ABS").replace("!=", "<>").replace("not", "NOT")
    

    if "EXISTS" not in converted_equation and len(re.findall(r'\b(AND|OR)\b', converted_equation)) > 1 and not match and "LIKE" not in converted_equation:
        
        segments = re.split(r'(\bAND\b|\bOR\b)', converted_equation)
        
        new_segments = []
        for i in range(0, len(segments)-1, 2):
            new_segments.append('(' + segments[i].strip() + ')')
            new_segments.append(segments[i+1].strip())
        new_segments.append('(' + segments[-1].strip() + ')')
        
        converted_equation = ' '.join(new_segments)
        
        result =  converted_equation
    elif match or 'SUM' in converted_equation or 'not(not' in equation:
        
        result =  converted_equation
    else:
        
        result =  add_brackets(converted_equation)
     
    if ('MIN' in result or 'MAX' in result) and 'MIN ( MAX' not in result and 'MAX ( MIN' not in result :
        
        result = reformat_min_max_expression(result).upper()
    
    
    return result
        

# Load the workbook and worksheet
file_path1 = f"i:/Validations- in progress.xlsx"
workbook = openpyxl.load_workbook(file_path1)
sheet = workbook["Sheet1"]
z_1_6_sheet = workbook["Z 1-6"]
keytext_sheet = workbook["KeyText"]
metrics_dict = build_metrics_dict(workbook)
z_1_6_data = create_z_1_6_dict(z_1_6_sheet)
keytext_data = create_KeyText_dict(keytext_sheet)


row = 1
while sheet.cell(row=row, column=3).value:
    cell_a_value = sheet.cell(row=row, column=1).value
    cell_b_value = sheet.cell(row=row, column=2).value
    cell_c_value = sheet.cell(row=row, column=3).value
    cell_b_value = cell_b_value.replace("dv: emptySequence()", "dv: emptySequence") if cell_b_value else cell_b_value
    cell_c_value = cell_c_value.replace("dv: emptySequence()", "dv: emptySequence") if cell_c_value else cell_c_value
    c_values_from_structure_c = get_c_values_from_structure_c(cell_a_value)
    r_values_from_structure_c = get_r_values_from_structure_c(cell_a_value)

    # Store the converted values temporarily
    converted_values = []
    for c_value in c_values_from_structure_c:
        for r_value in r_values_from_structure_c:
            converted_value = convert_if(cell_c_value, c_value, r_value, workbook, metrics_dict)
            converted_value = re.sub(r'\(\s*', '( ',  converted_value)
            converted_value = re.sub(r'\s*\)', ' )',  converted_value)
            converted_value = re.sub(r'NOT\s*\(', 'NOT (', converted_value)

            converted_values.append(converted_value)

    for idx_offset, converted_value in enumerate(converted_values):
        # Fetch the current value in column 5 for the specific row
        current_value_col5 = sheet.cell(row=row + idx_offset, column=5).value or ""
        sheet.cell(row=row + idx_offset, column=5).value = converted_value

        t_values_found = re.findall(r'[A-Z]\.\d{2}\.\d{2}\.\d{2}', converted_value)
        unique_t_values = list(set(t_values_found))
    
        for col_offset, t_value in enumerate(unique_t_values, start=6):
            sheet.cell(row=row + idx_offset, column=col_offset).value = t_value

        s06_t_value = next((t for t in unique_t_values if t.startswith('S.06.02')), None)
        convert_if_value = convert_if(cell_b_value, c_value, r_value, workbook, metrics_dict) if cell_b_value else ""

        if s06_t_value:
            additional_text_format = "TEXT( #REPPER, '#COMPANY', '{}', 'SCRC0150', '', '#FISCVARNT', '', '', '', '', '', '', 'ELEM' ) <> 'AGG_ONLY'"
            additional_text = additional_text_format.format(s06_t_value)
            
            if additional_text not in current_value_col5:
                if cell_b_value:
                    cell_value = "( " + convert_if_value + " ) AND " + additional_text
                else:
                    cell_value = additional_text
            else:
                cell_value = convert_if_value
        else:
            cell_value = convert_if_value

        sheet.cell(row=row + idx_offset, column=4).value = cell_value

        if idx_offset < len(converted_values) - 1:
            sheet.insert_rows(row + idx_offset + 1)

    row += len(converted_values)

workbook.save(file_path1)

